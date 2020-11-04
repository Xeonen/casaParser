# -*- coding: utf-8 -*-
"""
@author: Xeonen
"""

from io import StringIO
import pandas as pd
from glob import glob
import xlrd

from copy import deepcopy
from random import randint, randrange
from openpyxl import load_workbook
from openpyxl.styles.fills import PatternFill
from openpyxl.worksheet.copier import WorksheetCopy 

from tqdm import tqdm

from django.conf import settings

class excelProcedure():

    varData = """ID;motil;motilCond;motilCondMax;live;liveCond;liveCondMax;dense;denseCond;denseCondMax;number;numberCond;numberCondMax;head;headCond;other;otherCond;total;totalCond
1;B32;5;95;B33;40;99;A49;FALSE;FALSE;D19;1;5;A49;15;A49;15;D20;30
2;D19;40;95;D20;50;99;D21;5;25;D22;5;15;A49;15;A49;15;D23;30
3;D18;30;95;D20;50;99;D21;100;800;D22;40;150;D23;30;D24;20;D25;40
4;D19;40;95;D21;50;99;D22;15;30;D23;15;25;D24;15;D25;20;D26;30
5;D19;30;95;D21;50;99;D22;100;400;D23;50;100;D24;25;D25;20;D26;40
6;D18;35;95;A49;FALSE;99;A49;FALSE;1000;A49;FALSE;1000;A49;100;A49;100;A49;100"""

    varDF = pd.read_csv(StringIO(varData), sep=";", encoding="UTF-8")
    trash = "A49"
    changeDF = pd.DataFrame(
        { "01": ["B31", "A38"],
          "02": ["B35", "A42"],
          "03": ["B36", "A43"],
          "04": ["B37", "A44"],
          "05": ["B37", "A44"],
          "06": [trash, trash]
            }
        )
    
    checkList = ["motil", "live", "dense", "number"]
    checkListMal = ["head", "other", "total"]

    
    def __init__(self, source, dataset, testType, payetVol):
        self.varDF = excelProcedure.varDF.loc[excelProcedure.varDF["ID"] == testType]
        self.varDF.reset_index(drop=True, inplace=True)
        self.changeDF = excelProcedure.changeDF.loc[:, "0"+str(testType)]
        self.wb = load_workbook(source)
        self.source = self.wb["sourceSheet"]
        self.df = pd.read_excel(dataset)  #, encoding="UTF-8")
        self.fileList = glob("media/data/*.xls")
        
        self.trash = excelProcedure.trash
        self.payetVol = payetVol
        self.checkListMal = excelProcedure.checkListMal
        self.checkList = excelProcedure.checkList
        self. yellowFill = PatternFill(start_color='FFF033', end_color='FFF033', fill_type='solid')
        

    def parse_dense(self, denseStr):
        denseList = denseStr.split(" ")[0].split(",")
        dense = round(float(f"{denseList[0]}.{denseList[1]}"))
        return(dense)

        
    def get_casa(self, fileName):
        casaDict = {"motil": 0, "live": 0, "dense": 0}
        workbook  = xlrd.open_workbook(fileName)
        ws = workbook.sheet_by_index(0)
        try:
            motilRaw = ws.cell_value(26, 23)
            motil = round(motilRaw)
            dense = self.parse_dense(ws.cell_value(14, 5))
            if motil > 50:
                rnd = randint(1,9)
                live = motil + rnd
            elif motil >= 40:
                rnd = randint(1,9)
                live = 50+rnd
            else:
                rnd = randint(1,9)
                live = motil + rnd
        except Exception as e:
            motil = 0
            live = 0
            dense = 0

        casaDict["motil"] = int(round(motil, 0))
        casaDict["live"] = int(round(live, 0))
        casaDict["dense"] = int(round(dense, 0))
        casaDict["number"] = int(round(dense*motil*self.payetVol*0.01))
        
        return(casaDict)
    
    
    def gen_disfunc(self, motil, minVal, headMax, otherMax):
        multiplier = (randint(1, 100) + 2*minVal - motil)/(100+minVal)

        head = int(round(headMax*multiplier, 0))
        other = int(round(otherMax*multiplier, 0))

        funcDict = {"head": head, "other": other, "total": head+other}

        return(funcDict)
    
    def fillForm(self):
        fileList = deepcopy(self.fileList)        
        for ID in tqdm(self.df.ID):
            try:
                ID = int(ID)
            except:
                continue
            
            i = self.df[self.df["ID"] == ID].index[0]
            grade = True
            try:
                sheetIndex = int(self.df.loc[i, "ID"])
            except:
                continue
            sheetName = f"Sheet{sheetIndex}"
            
            fileName = str()
            sheetID = str(ID)
            sheetID = "0"*(3-len(sheetID))+sheetID
            for fID, file in enumerate(fileList):
                if settings.DEBUG:
                    file = file.split("\\")[-1]
                else:
                    file = file.split("/")[-1]

                if file.startswith(sheetID):
                    fileName = fileList.pop(fID)
                    break
            if len(fileName) < 1: print(f"Error in {sheetName}")
            

                       
            self.wb.create_sheet(sheetName)
            
            WorksheetCopy(self.source, self.wb[sheetName]).copy_worksheet()
            

            self.wb[sheetName]["B3"].value = sheetIndex
            self.wb[sheetName]["B9"].value = self.df.loc[i, "earID"]
            self.wb[sheetName]["B10"].value = self.df.loc[i, "race"]
            self.wb[sheetName]["B11"].value = self.df.loc[i, "name"]
            self.wb[sheetName]["B12"].value = self.df.loc[i, "prodDate"]
            
            try:
                self.wb[sheetName]["B13"].value = self.df.loc[i, "lotID"]
            except:
                self.wb[sheetName]["B13"].value = ""
            
            self.wb[sheetName]["B14"].value = self.df.loc[i, "sample1"]
            
            casaDict = self.get_casa(fileName)
            
            motil = casaDict['motil']
            dense = casaDict['dense']
            live = casaDict['live']           
            number = casaDict['number']


            for checkObj in self.checkList:
                checkObjCond = checkObj.lower() + "Cond"
                checkObjCondMax = checkObjCond+"Max"
                cond = self.varDF.loc[0, checkObjCond]
                condMax = self.varDF.loc[0, checkObjCondMax]
                loc = self.varDF.loc[0, checkObj]
                val = casaDict[checkObj]


                if cond != "FALSE":
                    cond = int(cond)
                    condMax = int(condMax)
                    if val < cond:
                        self.wb[sheetName][loc].fill = self.yellowFill
                        grade = False
                    elif val > condMax:
                        val = int(round(condMax*randint(90, 95)/100, 0))

                    self.wb[sheetName][loc].value = val


            headMax = int(self.varDF.loc[0, "headCond"])
            otherMax = int(self.varDF.loc[0, "otherCond"])
            minVal = int(self.varDF.loc[0, "motilCond"])

            funcDict = self.gen_disfunc(motil, minVal, headMax, otherMax)
            
            for key in funcDict.keys():
                checkObjCond = key.lower() + "Cond"
                loc = self.varDF.loc[0, key]
                val = funcDict[key]
                self.wb[sheetName][loc].value = funcDict[key]
                if val > int(self.varDF.loc[0, checkObjCond]):
                    self.wb[sheetName][loc].fill = self.yellowFill

                
            
            if grade == True:
                resultText = "Yukarıda üretici/ithalatçı bilgileri ve spermatolojik sonuçları verilen sperma Suni Tohumlama uygulamalarında kullanılabilir özelliktedir."
            else:
                resultText = "Yukarıda üretici/ithalatçı bilgileri ve spermatolojik sonuçları verilen sperma Suni Tohumlama uygulamalarında kullanılabilir özellikte değildir."
                
            amountLoc = self.changeDF.iloc[0]           
            resultLoc = self.changeDF.iloc[1]    
            
            self.wb[sheetName][amountLoc].value = self.df.loc[i, "amount"]
            self.wb[sheetName][resultLoc].value = resultText
            self.wb[sheetName][self.trash].value = ""
            
        self.wb.save("media/casaRapor.xlsx")
        self.wb.close()

            
                        
                
            
                
                
                





        
        
        
        




      
# ep = excelProcedure("source.xlsx", "dataset.xlsx", 1, 0.25)
# ep.fillForm()




