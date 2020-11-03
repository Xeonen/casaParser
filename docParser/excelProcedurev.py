# -*- coding: utf-8 -*-
"""
@author: Xeonen
"""

from copy import copy
import pandas as pd
from glob import glob
import xlrd

from copy import deepcopy
from random import randint, randrange
from openpyxl import load_workbook
from openpyxl.styles.fills import PatternFill
from openpyxl.worksheet.copier import WorksheetCopy 

from tqdm import tqdm

class excelProcedure():
    
    varDF = pd.read_csv("varData.csv", sep=";", encoding="UTF-8")
    trash = "A49"
    changeDF = pd.DataFrame(
        { "01": ["B31", "A38"],
          "02": ["B35", "A42"],
          "03": ["B36", "A38"],
          "04": ["B37", "A38"],
          "05": ["B37", "A39"],
          "06": [trash, "D18"]        
            }
        )
    
    checkList = ["motil", "live", "dense"]

    
    def __init__(self, source, dataset, testType, payetVol):
        self.varDF = excelProcedure.varDF.loc[excelProcedure.varDF["ID"] == testType]
        self.varDF.reset_index(drop=True, inplace=True)
        self.changeDF = excelProcedure.changeDF.loc[:, "0"+str(testType)]
        self.wb = load_workbook(source)
        self.source = self.wb["sourceSheet"]
        self.df = pd.read_excel(dataset)  #, encoding="UTF-8")
        self.fileList = glob("data/*.xls")
        
        self.trash = excelProcedure.trash
        self.payetVol = payetVol
        self.checkList = excelProcedure.checkList
        self. yellowFill = PatternFill(start_color='FFF033', end_color='FFF033', fill_type='solid')
        

    def parse_dense(self, denseStr):
        denseList = denseStr.split(" ")[0].split(",")
        dense = round(float(f"{denseList[0]}.{denseList[1]}"))
        return(dense)

        
    def get_casa(self, fileName):
        casaDict = {"motil": 0, "live": 0, "dense": 0}
        workbook  = xlrd.open_workbook(fileName)
        ws = workbook.sheet_by_index(0)
        try:
            motilRaw = ws.cell_value(26, 23)
            motil = round(motilRaw)
            dense =  self.parse_dense(ws.cell_value(14, 5))            
            if motil > 50:
                rnd = randint(1,9)
                live = motil + rnd
            elif motil >= 40:
                rnd = randint(1,9)
                live = 50+rnd
            else:
                rnd = randint(1,9)
                live = motil + rnd
        except Exception as e:
            motil = 0
            live = 0
            dense = 0
        
        casaDict["motil"] = motil
        casaDict["live"] = live
        casaDict["dense"] = dense
        casaDict["number"] = dense*motil*self.payetVol*0.01
        
        return(casaDict)
    
    
    def gen_disfunc(self, grade):
    
        above_30 = True
        if not grade:
            head = randint(1,15) + randint(1, 4)
            other = randint(1,20) + randint(1, 4)
            if head+other < 29:
                above_30 = False        
        else:
            while above_30:
                head = randint(1,10) 
                other = randint(1,12)
                if head+other < 29:
                    above_30 = False
                    
        funcDict = {"head": head, "other": other, "total": head+other}
        return(funcDict)
    
    def fillForm(self):
        fileList = deepcopy(self.fileList)        
        for ID in tqdm(self.df.ID):
            try:
                ID = int(ID)
            except:
                continue
            
            i = self.df[self.df["ID"] == ID].index[0]
            grade = True
            try:
                sheetIndex = int(self.df.loc[i, "ID"])
            except:
                continue
            sheetName = f"Sheet{sheetIndex}"
            
            fileName = str()
            sheetID = str(ID)
            sheetID = "0"*(3-len(sheetID))+sheetID
            for fID, file in enumerate(fileList):
                file = file.split("\\")[-1]
                if file.startswith(sheetID):
                    fileName = fileList.pop(fID)
                    break
            if len(fileName) < 1: print(f"Error in {sheetName}")
            

                       
            self.wb.create_sheet(sheetName)
            
            WorksheetCopy(self.source, self.wb[sheetName]).copy_worksheet()
            

            self.wb[sheetName]["B3"].value = sheetIndex
            self.wb[sheetName]["B9"].value = self.df.loc[i, "earID"]
            self.wb[sheetName]["B10"].value = self.df.loc[i, "race"]
            self.wb[sheetName]["B11"].value = self.df.loc[i, "name"]
            self.wb[sheetName]["B12"].value = self.df.loc[i, "prodDate"]
            
            try:
                self.wb[sheetName]["B13"].value = self.df.loc[i, "lotID"]
            except:
                self.wb[sheetName]["B13"].value = ""
            
            self.wb[sheetName]["B14"].value = self.df.loc[i, "sample1"]
            
            casaDict = self.get_casa(fileName)
            
            motil = casaDict['motil']
            dense = casaDict['dense']
            live = casaDict['live']           
            number = casaDict['number']
            

            
            for checkObj in self.checkList:
                

                checkObjCond = checkObj.lower() + "Cond"
                
                cond = self.varDF.loc[0, checkObjCond]  

                    
                loc = self.varDF.loc[0, checkObj]
                val = casaDict[checkObj]
                    
                self.wb[sheetName][loc].value = val

                if cond != "FALSE":
                    cond = int(cond)
                    if val < cond:
                        self.wb[sheetName][loc].fill = self.yellowFill
                        grade = False

                        
            funcDict = self.gen_disfunc(grade)
            
            for key in funcDict.keys():
                loc = self.varDF.loc[0, key]
                self.wb[sheetName][loc].value = funcDict[key]
                
            
            if grade == True:
                resultText = "Yukarıda üretici/ithalatçı bilgileri ve spermatolojik sonuçları verilen sperma Suni Tohumlama uygulamalarında kullanılabilir özelliktedir."
            else:
                resultText = "Yukarıda üretici/ithalatçı bilgileri ve spermatolojik sonuçları verilen sperma Suni Tohumlama uygulamalarında kullanılabilir özellikte değildir."
                
            amountLoc = self.changeDF.iloc[0]           
            resultLoc = self.changeDF.iloc[1]    
            
            self.wb[sheetName][amountLoc].value = self.df.loc[i, "amount"]
            self.wb[sheetName][resultLoc].value = resultText
            self.wb[sheetName][self.trash].value = ""
            
        self.wb.save("neoResult.xlsx")
        self.wb.close()

            
                        
                
            
                
                
                





        
        
        
        




      
ep = excelProcedure("source.xlsx", "dataset.xlsx", 1, 0.25)
ep.fillForm()




